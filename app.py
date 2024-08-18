import streamlit as st
from ultralytics import YOLO
import numpy as np
from PIL import Image
import cv2
import matplotlib.pyplot as plt
from collections import defaultdict
from fpdf import FPDF
import pandas as pd
from openpyxl import load_workbook
import tempfile
import os

# YOLO model initialization
model = YOLO('model.pt')

# App Layout
st.set_page_config(page_title="Traffic Pattern Analysis", layout="wide")

# Use session state to show effects only once
if 'show_effects' not in st.session_state:
    st.session_state.show_effects = True

if st.session_state.show_effects:
    st.balloons()
    st.session_state.show_effects = False

# Initialize session state for persistent image and detection data
if 'uploaded_image' not in st.session_state:
    st.session_state.uploaded_image = None
if 'detected_image' not in st.session_state:
    st.session_state.detected_image = None
if 'detection_data' not in st.session_state:
    st.session_state.detection_data = defaultdict(lambda: {'Count': 0, 'Confidence': []})
if 'captured_video' not in st.session_state:
    st.session_state.captured_video = None

# Custom CSS for Traffic Light Theme
st.markdown("""
    <style>
    /* Set the entire page background color to light white */
    .stApp {
        background-color: #f9f9f9;
        color: #000000;
    }

    /* Sidebar background and text color */
    .css-1d391kg {
        background-color: #f9f9f9 !important;
        color: #000000 !important;
    }

    /* Sidebar title color */
    .css-1d391kg h1 {
        color: #ffffff !important;
    }

    /* Sidebar titles and labels */
    .css-1d391kg h2, .css-1d391kg label {
        color: #ffffff !important;
    }
    .css-1n543e5 {
        font-size: 24px; /* Adjust this value for your desired size */
        font-weight: bold;
        color: #ffffff;
    }
    .sidebar-title {
        color: white;
        font-size: 24px; /* Adjust font size if needed */
        font-weight: bold;
    }
    .sidebar-radio label {
        color: white; /* This will make the radio labels white */
        font-size: 18px; /* Adjust font size if needed */
    }

    /* General text color */
    body {
        color: #000000;
    }
    
    /* Header styling */
    h1, h2, h3 {
        font-family: 'Arial', sans-serif;
        font-weight: bold;
        color: #000000;
    }

    /* Buttons */
    .stButton>button {
        background-color: #ffffff;
        color: #000000;
        font-size: 20px;
        border: 2px solid #000000;
        border-radius: 10px;
    }

    .stButton>button:hover {
        background-color: #000000;
        color: #ffffff;
    }

    /* Form inputs */
    .stTextInput>div>div>input {
        border-radius: 10px;
        border: 2px solid #000000;
    }

    .stTextArea>div>div>textarea {
        border-radius: 10px;
        border: 2px solid #000000;
    }
            
    /* Header styling for logo */
    .header {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding: 10px 0;
    }
            
    .dashboard-header {
        color: white; /* Set the text color to white */
        font-size: 32px; /* Adjust the font size if needed */
    }

    /* Sliders */
    .stSlider>div>div>div>div {
        background-color: #000000;
    }

    /* Tables */
    .stDataFrame>div {
        border-radius: 10px;
        border: 2px solid #000000;
    }

    /* Titles */
    .title {
        font-size: 32px;
        color: #000000;
    }

    /* Chart styling */
    .stPlotlyChart {
        margin-top: 20px;
    }
            
    /* Button styling for Download Report to match Start and Stop buttons */
    .stDownloadButton > button {
        background-color: #ffffff;
        color: #000000;
        font-size: 18px;
        border: 2px solid #000000;
        border-radius: 10px;
        padding: 10px 20px;
    }

    .stDownloadButton > button:hover {
        background-color: #000000;
        color: #ffffff;
    }
            
    /* Form text input and area styling */
    .stTextInput > div > div > input, .stTextArea > div > div > textarea {
        color: #000000 !important; /* Black text color */
        background-color: #ffffff !important; 
    }
            
    /* Label styling */
    label {
        color: #000000 !important; /* Black text color */
    }
            
     /* Specific Radio button options styling for the feedback form */
    div[role='radiogroup'] > label > div[data-testid='stMarkdownContainer'] {
        color: #000000 !important; 
    }

    /* Ensure other radio buttons (not in the feedback form) stay in white */
    div[role='radiogroup'] > label > div:not([data-testid='stMarkdownContainer']) {
        color: #ffffff !important; 
    }
    
    .logo-container {
        position: absolute;
        top: 0;
        right: 0;
        padding: 10px;
    }
            
    .logo-container img {
        width: 150px;  /* Adjust the width as needed */
    }
    
    .white-arrow {
        color: white; /* Set the color to white */
        font-size: 30px; /* Increase the font size */
        margin: 0; 
        padding: 0; 
    }
            
    </style>
    """, unsafe_allow_html=True)

st.markdown(f"""
    <div class="header">
        <h1>Traffic Pattern Analysis 🚦</h1>
    </div>
    """, unsafe_allow_html=True)

# Sidebar for Navigation
st.sidebar.markdown('<h1 class="sidebar-title">Dashboard</h1>', unsafe_allow_html=True)
st.sidebar.markdown('<p class="white-arrow">🔽</p>', unsafe_allow_html=True)
options = st.sidebar.radio(
    '',
    ["Select an option", "Real-Time Detection", "Image Detection", "Graphical Visualization", "User Feedback", "Developer Details", "Download Report"]
)
# Display the Title


# Real-Time Detection Feature
if options == "Real-Time Detection":
    st.header("Real-Time Detection 📹")

    stframe = st.empty()
    cap = cv2.VideoCapture(0)

    start_button = st.button('Start', key='start_button')
    stop_button = st.button('Stop', key='stop_button')

    if stop_button:
        st.markdown("<p style='font-size:20px;'>Stopping video feed...</p>", unsafe_allow_html=True)
        cap.release()
        cv2.destroyAllWindows()
        stframe.empty()
    elif start_button:
        st.markdown("<p style='font-size:20px;'>Starting video feed...</p>", unsafe_allow_html=True)
        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                st.markdown("<p style='font-size:20px;'>Failed to capture image.</p>", unsafe_allow_html=True)
                break

            # Perform detection
            results = model.predict(frame)

            # Process results and render
            annotated_frame = frame
            for result in results:
                annotated_frame = result.plot()  # Get annotated frame
                for obj in result.boxes:  # Iterate over detected objects
                    class_name = model.names[int(obj.cls)]
                    confidence = obj.conf
                    st.session_state.detection_data[class_name]['Count'] += 1
                    st.session_state.detection_data[class_name]['Confidence'].append(confidence)

            rgb_frame = cv2.cvtColor(annotated_frame, cv2.COLOR_BGR2RGB)
            stframe.image(rgb_frame, channels='RGB', use_column_width=True)

        cap.release()
        cv2.destroyAllWindows()

# Image Detection Feature
elif options == "Image Detection":
    st.header("Image Detection 📸")

    uploaded_image = st.file_uploader("Upload an Image", type=["jpg", "jpeg", "png"])

    if uploaded_image is not None:
        st.session_state.uploaded_image = uploaded_image

    if st.session_state.uploaded_image:
        image = Image.open(st.session_state.uploaded_image)
        st.image(image, caption='Uploaded Image', use_column_width=True)
        
        if st.button('Detect', key='detect_button'):
            # Convert image to BGR for YOLO
            image_np = np.array(image)
            image_bgr = cv2.cvtColor(image_np, cv2.COLOR_RGB2BGR)
            
            # Perform detection
            results = model.predict(image_bgr)

            # Process results and render
            annotated_image = image_np
            for result in results:
                annotated_image = result.plot()  # Get annotated image
                for obj in result.boxes:  # Iterate over detected objects
                    class_name = model.names[int(obj.cls)]
                    confidence = obj.conf
                    st.session_state.detection_data[class_name]['Count'] += 1
                    st.session_state.detection_data[class_name]['Confidence'].append(confidence)

            st.session_state.detected_image = Image.fromarray(cv2.cvtColor(annotated_image, cv2.COLOR_BGR2RGB))
            st.image(st.session_state.detected_image, caption='Detected Image', use_column_width=True)

# Graphical Visualization Feature
elif options == "Graphical Visualization":
    st.header("Graphical Visualization 📊")

    if st.session_state.detection_data:
        class_names = list(st.session_state.detection_data.keys())
        counts = [st.session_state.detection_data[class_name]['Count'] for class_name in class_names]

        if counts:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.bar(class_names, counts, color='green')
            ax.set_xlabel('Object Classes')
            ax.set_ylabel('Count')
            ax.set_title('Detected Objects Count')
            plt.xticks(rotation=45, ha='right')  # Rotate x-axis labels for better readability

            st.pyplot(fig)

            # Display counts
            st.markdown("<p style='font-size:20px;'>Detected Objects Count:</p>", unsafe_allow_html=True)
            for class_name in class_names:
                st.write(f"**{class_name}:** {st.session_state.detection_data[class_name]['Count']}")
        else:
            st.markdown("<p style='font-size:20px;'>No objects detected to display.</p>", unsafe_allow_html=True)
    else:
        st.markdown("<p style='font-size:20px;'>No detection data available.</p>", unsafe_allow_html=True)

# User Feedback Feature

elif options == "User Feedback":
    file_path = r"C:\\Users\\Rithu\\OneDrive\\userfeedback.xlsx"
    st.header("User Feedback 💬")
    
    with st.form(key='feedback_form'):
        name = st.text_input("Your Name")
        email = st.text_input("Your Email")
        feedback = st.text_area("Please provide your feedback about the app:")

        st.markdown("**Rate the App:**")
        rating = st.radio(
            "Select a rating:",
            options=["🟢 (5 stars)", "🟡 (4 stars)", "🔴 (3 stars)"],
            index=0
        )
        
        submit_button = st.form_submit_button(label='Submit')
    
    if submit_button:
        rating_map = {
            "🟢 (5 stars)": 5,
            "🟡 (4 stars)": 4,
            "🔴 (3 stars)": 3
        }
        selected_rating = rating_map.get(rating, None)
        
        if selected_rating is None:
            st.warning("Please select a rating before submitting.")
        else:
            st.success("Thank you for your feedback!")
            # Display the feedback
            st.write(f"**Name:** {name}")
            st.write(f"**Email:** {email}")
            st.write(f"**Feedback:** {feedback}")
            st.write(f"**Rating:** {selected_rating} stars")
            feedback_data = {
                "Name": [name],
                "Email": [email],
                "Rating": [selected_rating],
                "Feedback": [feedback]
            }
            df = pd.DataFrame(feedback_data)
            
            try:
                # Try to load existing file
                existing_data = pd.read_excel(file_path)
                df = pd.concat([existing_data, df], ignore_index=True)
            except FileNotFoundError:
                # If file doesn't exist, this will be the first entry
                pass
            
            df.to_excel(file_path, index=False)
# Developer Details Feature
elif options == "Developer Details":
    st.header("Developer Details 🧑‍💻")
    
    st.markdown("<p style='font-size:20px;'><strong>Name:</strong> Ramya R</p>", unsafe_allow_html=True)
    st.markdown("<p style='font-size:20px;'><strong>Project Title:</strong> Traffic Pattern Analysis using Machine Learning</p>", unsafe_allow_html=True)
    st.markdown("<p style='font-size:20px;'><strong>College:</strong> Panimalar Engineering College</p>", unsafe_allow_html=True)
    st.markdown("<p style='font-size:20px;'><strong>Contact:</strong> ramyajaramesh@gmail.com </p>", unsafe_allow_html=True)

# Download Report Feature
elif options == "Download Report":
    st.header("Download Report 📥")

    if st.session_state.uploaded_image or st.session_state.detected_image or st.session_state.detection_data:
        # Generate PDF report
        pdf = FPDF()
        pdf.add_page()

        # Title
        pdf.set_font("Arial", size=20)
        pdf.cell(200, 10, txt="Traffic Pattern Analysis Report", ln=True, align='C')

        # Uploaded Image
        if st.session_state.uploaded_image:
            uploaded_img = Image.open(st.session_state.uploaded_image)
            uploaded_img_path = tempfile.mktemp(suffix=".jpg")
            uploaded_img.save(uploaded_img_path)
            pdf.image(uploaded_img_path, x=10, y=30, w=180)
        
        # Detected Image
        if st.session_state.detected_image:
            detected_img = st.session_state.detected_image
            detected_img_path = tempfile.mktemp(suffix=".jpg")
            detected_img.save(detected_img_path)
            pdf.add_page()
            pdf.image(detected_img_path, x=10, y=30, w=180)

        # Graphical Visualization
        if st.session_state.detection_data:
            class_names = list(st.session_state.detection_data.keys())
            counts = [st.session_state.detection_data[class_name]['Count'] for class_name in class_names]

            if counts:
                fig, ax = plt.subplots(figsize=(10, 6))
                ax.bar(class_names, counts, color='green')
                ax.set_xlabel('Object Classes')
                ax.set_ylabel('Count')
                ax.set_title('Detected Objects Count')
                plt.xticks(rotation=45, ha='right')

                plt.tight_layout()
                graph_path = tempfile.mktemp(suffix=".png")
                plt.savefig(graph_path)
                plt.close(fig)  # Close the plot to release the file
                pdf.add_page()
                pdf.image(graph_path, x=10, y=30, w=180)
        
        # Counts
        if st.session_state.detection_data:
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt="Detected Objects Count", ln=True)

            for class_name in class_names:
                count = st.session_state.detection_data[class_name]['Count']
                pdf.cell(200, 10, txt=f"{class_name}: {count}", ln=True)

        # Output PDF to file
        pdf_path = tempfile.mktemp(suffix=".pdf")
        pdf.output(pdf_path)

        # Provide download link
        with open(pdf_path, "rb") as f:
            st.download_button(
                label="Download Report",
                data=f,
                file_name="traffic_pattern_analysis_report.pdf",
                mime="application/pdf"
            )

        # Ensure the file is removed after download
        os.remove(pdf_path)
        if os.path.exists(uploaded_img_path):
            os.remove(uploaded_img_path)
        if os.path.exists(detected_img_path):
            os.remove(detected_img_path)
        if os.path.exists(graph_path):
            os.remove(graph_path)
    else:
        st.markdown("<p style='font-size:20px;'>No data available for the report.</p>", unsafe_allow_html=True)
